import sqlite3
import openpyxl

# import matplotlib.pyplot as plt
import pandas as pd

# from matplotlib.font_manager import FontProperties
from collections import OrderedDict
import sys

# import pandas as pd

sys.stdout.reconfigure(encoding="utf-8")


# 将从数据库中查询返回的结果储存为字典的形式，键名为列名
def select_all_dict():
    # 连接数据库
    conn = sqlite3.connect("database.db")
    c = conn.cursor()
    # 检查目标数据表是否存在
    c.execute("select name from sqlite_master where type ='table' and name = 'elastic'")
    table_exit = c.fetchone()
    if table_exit:
        print("elastic数据表已存在，数据表中的所有信息为：")

        c.execute("SELECT COUNT(*) FROM elastic;")
        rows = c.fetchone()[0]
        if rows > 0:
            print("表中有实际数据")
        else:
            list = openpyxl.load_workbook("data.xlsx")
            datalist = list.active
            data_insert = """INSERT INTO elastic(id,index_name,shard,prirep,state,docs,store,ip,node) VALUES (?,?,?,?,?,
            ?,?,?,?); """
            # 遍历打开的excel中的每一行数据，通过迭代的方法将每一行的数据对应地写入到数据库中
            for row in datalist.iter_rows(
                min_row=2, max_col=9, max_row=(datalist.max_row - 1)
            ):
                # 使excel各行数据成为迭代器
                cargo = [cell.value for cell in row]  # 敲黑板！！使每行中单元格成为迭代器
                # 将“id”列中的值转换为整数类型
                cargo[0] = int(cargo[0]) if cargo[0] is not None else None
                # ...
                c.execute(data_insert, cargo)  # 敲黑板！写入一行数据到数据库中表mylist
            conn.commit()
            print("数据写入成功")

    if not table_exit:
        # 若检查的数据表不存在则创建数据表
        c.execute(
            """create table elastic (id integer not null primary key autoincrement ,index_name text,shard text,prirep text,
        state text,docs text,store text,ip text,node text ); """
        )
        conn.commit()
        print("elastic数据表创建成功")

        # 导入excel表格并将其中的数据写入到数据库中
        list = openpyxl.load_workbook("data.xlsx")
        datalist = list.active
        data_insert = """INSERT INTO elastic(id,index_name,shard,prirep,state,docs,store,ip,node) VALUES (?,?,?,?,?,
        ?,?,?,?); """
        # 遍历打开的excel中的每一行数据，通过迭代的方法将每一行的数据对应地写入到数据库中
        for row in datalist.iter_rows(
            min_row=2, max_col=9, max_row=(datalist.max_row - 1)
        ):
            # 使excel各行数据成为迭代器
            cargo = [cell.value for cell in row]  # 敲黑板！！使每行中单元格成为迭代器
            # 将“id”列中的值转换为整数类型
            cargo[0] = int(cargo[0]) if cargo[0] is not None else None
            # ...
            c.execute(data_insert, cargo)  # 敲黑板！写入一行数据到数据库中表mylist
        conn.commit()
        print("数据写入成功")

    # 删除数据表中的最后一行
    # c.execute("delete from elastic where id = 19")
    # conn.commit()

    # 查询除了主键id之外所有的数据
    cursor = c.execute(
        "select index_name,shard,prirep,state,docs,store,ip,node from elastic order by node asc"
    )
    columns = [i[0] for i in cursor.description]
    results = []
    for row in cursor.fetchall():
        row_dict = {}
        for i, value in enumerate(row):
            row_dict[columns[i]] = value
        results.append(row_dict)
    conn.close()
    results1 = []

    # 修改数据库返回的键的值
    for item in results:
        item = OrderedDict(item)
        item["index"] = item.pop("index_name")
        item.move_to_end("index", last=False)
        item = dict(item)
        results1.append(item)

    return results1


# 获取所点击选取节点的信息
def select_first_shard_node(str):
    # 连接数据库
    conn = sqlite3.connect("database.db")
    c = conn.cursor()

    # 查询约束条件下的信息
    query = "select index_name,shard,prirep,state,docs,store,ip,node from elastic where index_name = ?  and prirep = ? and shard = ?"
    list_params = str.split("&")
    params = tuple(list_params)
    # excute函数的第二个参数要封装在一个元组中才可以正确进行传递
    cursor = c.execute(query, params)

    columns = [i[0] for i in cursor.description]
    results = []
    for row in cursor.fetchall():
        row_dict = {}
        for i, value in enumerate(row):
            row_dict[columns[i]] = value
        results.append(row_dict)
    conn.close()
    results1 = []

    # 修改数据库返回的键的值
    for item in results:
        item = OrderedDict(item)
        item["index"] = item.pop("index_name")
        item.move_to_end("index", last=False)
        item = dict(item)
        results1.append(item)

    return results1


# 获取所选index所有节点的信息
def select_first_index_node(condition_str):
    # 连接数据库
    conn = sqlite3.connect("database.db")
    c = conn.cursor()

    # 查询约束条件下的信息
    query = "select index_name,shard,prirep,state,docs,store,ip,node from elastic where index_name = ?"
    cursor = c.execute(query, (condition_str,))

    columns = [i[0] for i in cursor.description]
    results = []
    for row in cursor.fetchall():
        row_dict = {}
        for i, value in enumerate(row):
            row_dict[columns[i]] = value
        results.append(row_dict)
    conn.close()
    results1 = []

    # 修改数据库返回的键的值
    for item in results:
        item = OrderedDict(item)
        item["index"] = item.pop("index_name")
        item.move_to_end("index", last=False)
        item = dict(item)
        results1.append(item)

    return results1


# 第三问在20台物理机上分配虚拟机来实现最大化资源利用率的后端代码实现
def get_third_question_list():
    # 读取excel文件第一行作为表头不读入
    df1 = pd.read_excel(
        "C:/Software_textwork/Elasticsearch_module/api/BM&VM机器信息明细表.xlsx",
        sheet_name="虚拟机VM",
        header=0,
    )
    # 将读取的数据的列名设置为"name", "disk", "cpu" 和 "memory"
    df1.columns = ["name", "disk", "cpu", "memory"]

    # 将数据转换为列表的形式
    virtual_machines = df1.to_dict("records")

    count_v = len(virtual_machines)
    print(f"共有{count_v}台虚拟机")

    # 定义物理机和虚拟机的配置
    physical_machine = {"name": "BM-NEW", "cpu": 120, "memory": 512, "disk": 12516}
    #
    # # 初始化每台物理机的剩余资源
    # remaining_resources = [physical_machine.copy() for _ in range(20)]
    # count = len(remaining_resources)
    # print(f"共有{count}台物理机")

    # 从excel表中读取物理机的信息
    df2 = pd.read_excel(
        "C:/Software_textwork/Elasticsearch_module/api/BM&VM机器信息明细表.xlsx",
        sheet_name="物理机BM",
        header=0,
    )
    df2.columns = ["name", "disk", "cpu", "memory"]
    # 将数据转换为列表的形式
    remaining_resources = df2.to_dict("records")
    print(remaining_resources)

    # 分配虚拟机到物理机
    allocation_list = []
    allocation_result = {}
    for vm in virtual_machines:
        allocated = False
        for pm_index, pm in enumerate(remaining_resources):
            if all(vm[k] <= pm[k] for k in ("cpu", "memory", "disk")):
                # 分配虚拟机到物理机
                for k in ("cpu", "memory", "disk"):
                    pm[k] -= vm[k]
                # allocation_result[vm["name"]] = pm['name']
                allocation_list.append(
                    {
                        "physics": pm["name"],
                        "virtual": vm["name"],
                        "physics_storage": "12516gb",
                        "physics_cpu": "120",
                        "physics_memory": "512gb",
                        "virtual_storage": str(vm["disk"]) + "gb",
                        "virtual_cpu": str(vm["cpu"]),
                        "virtual_memory": str(vm["memory"]) + "gb",
                        "storage_Occupancy": str(
                            round(pm["cpu"] / physical_machine["cpu"] * 100, 2)
                        )
                        + "%",
                        "cpu_Occupancy": str(
                            round(pm["cpu"] / physical_machine["cpu"] * 100, 2)
                        )
                        + "%",
                        "memory_Occupancy": str(
                            round(pm["memory"] / physical_machine["memory"] * 100, 2)
                        )
                        + "%",
                    }
                )
                allocated = True
                break
        if not allocated:
            # 当前物理机没有足够的剩余资源可以分配给虚拟机，需要寻找其他未被分配的
            for pm_index, pm in enumerate(remaining_resources):
                for vm_name, vm_index in allocation_result.items():
                    if vm_index == pm_index:
                        continue
                    vm = next(
                        (x for x in virtual_machines if x["name"] == vm_name), None
                    )
                    if all(vm[k] <= pm[k] for k in ("cpu", "memory", "disk")):
                        for k in ("cpu", "memory", "disk"):
                            pm[k] -= vm[k]
                        # allocation_result[vm["name"]] = pm['name']
                        allocation_list.append(
                            {
                                "physics": pm["name"],
                                "virtual": vm["name"],
                                "physics_storage": "12516gb",
                                "physics_cpu": "120",
                                "physics_memory": "512gb",
                                "virtual_storage": str(vm["disk"]) + "gb",
                                "virtual_cpu": str(vm["cpu"]),
                                "virtual_memory": str(vm["memory"]) + "gb",
                                "storage_Occupancy": str(
                                    round(
                                        pm["disk"] / physical_machine["disk"] * 100, 2
                                    )
                                )
                                + "%",
                                "cpu_Occupancy": str(
                                    round(pm["cpu"] / physical_machine["cpu"] * 100, 2)
                                )
                                + "%",
                                "memory_Occupancy": str(
                                    round(
                                        pm["memory"] / physical_machine["memory"] * 100,
                                        2,
                                    )
                                )
                                + "%",
                            }
                        )
                        allocated = True
                        break
                if allocated:
                    break
            if not allocated:
                # 已有的物理机都分配满了，无法再创建新的物理机来分配该虚拟机
                print(
                    "No available physical machine to allocate VM: {}".format(
                        vm["name"]
                    )
                )

    # 循环遍历已分配的虚拟机列表，计算每台虚拟机所在物理机的最终的资源利用率
    for vm in allocation_list:
        pm = [pm for pm in remaining_resources if pm["name"] == vm["physics"]][0]
        vm["storage_Occupancy"] = (
            str(round((1 - pm["disk"] / physical_machine["disk"]) * 100, 2)) + "%"
        )
        vm["cpu_Occupancy"] = (
            str(round((1 - pm["cpu"] / physical_machine["cpu"]) * 100, 2)) + "%"
        )
        vm["memory_Occupancy"] = (
            str(round((1 - pm["memory"] / physical_machine["memory"]) * 100, 2)) + "%"
        )

    print(remaining_resources)
    count2 = len(remaining_resources)
    print(f"当前共有{count2}台物理机")

    sorted_allocated_vms = sorted(
        allocation_list, key=lambda x: int(x["physics"].split("-")[1])
    )

    return sorted_allocated_vms


def get_third_q_physics_node(params):
    # 更改参数中的键名'name'为'physics'
    value = params.pop("name")
    params["physics"] = value
    # 将get_third_question_list()返回值中满足参数中'physics'对应的的元素保存在列表中
    all_node_list = get_third_question_list()
    physics_node_list = [
        d for d in all_node_list if all(k in d and d[k] == v for k, v in params.items())
    ]

    print(physics_node_list)
    return physics_node_list


if __name__ == "__main__":
    # 连接数据库
    # conn = sqlite3.connect("database.db")
    # print("数据库连接成功")
    # # 创建光标对象
    # c = conn.cursor()
    # # 插入一条新的数据
    # # c.execute(
    # #     "insert into elastic (ip,node) values(?,?)", ("27.19.50.146", "data-node-07")
    # # )
    # # conn.commit()
    # # print("数据插入成功")
    # c.execute("select * from elastic")
    # rows = c.fetchall()
    # for row in rows:
    #     print(row)

    # # 执行 SQL 语句删除最后一条数据
    # # c.execute("DELETE FROM table_name WHERE rowid = (SELECT max(rowid) FROM table_name)")

    # conn.close()

    # # 检查目标数据表是否存在
    # c.execute("select name from sqlite_master where type ='table' and name = 'elastic'")
    # table_exit = c.fetchone()
    # if table_exit:
    #     print("elastic数据表已存在")

    # if not table_exit:
    #     # 若检查的数据表不存在则创建数据表
    #     c.execute(
    #         """create table elastic (id integer not null primary key autoincrement ,index_name text,shard text,prirep text,
    #     state text,docs text,store text,ip text,node text ); """
    #     )
    #     conn.commit()
    #     print("elastic数据表创建成功")

    #     # 导入excel表格并将其中的数据写入到数据库中
    #     list = openpyxl.load_workbook("api\data.xlsx")
    #     datalist = list.active
    #     data_insert = """INSERT INTO elastic(id,index_name,shard,prirep,state,docs,store,ip,node) VALUES (?,?,?,?,?,
    #     ?,?,?,?) """
    #     # 遍历打开的excel中的每一行数据，通过迭代的方法将每一行的数据对应地写入到数据库中
    #     for row in datalist.iter_rows(min_row=2, max_col=9, max_row=datalist.max_row):
    #         # 使excel各行数据成为迭代器
    #         cargo = [cell.value for cell in row]  # 敲黑板！！使每行中单元格成为迭代器
    #         # 将“id”列中的值转换为整数类型
    #         cargo[0] = int(cargo[0]) if cargo[0] is not None else None
    #         # ...
    #         c.execute(data_insert, cargo)  # 敲黑板！写入一行数据到数据库中表mylist
    #     conn.commit()
    #     print("数据写入成功")

    # conn.close()

    # 调用第三问的函数
    get_third_question_list()
